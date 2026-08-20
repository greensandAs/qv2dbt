"""Source-to-Target Mapping (STTM) generator: Excel workbook + YAML.

Produces the artefact a data team signs off before build:
  * Target Inventory  - every table, its layer/role and shape
  * Source Inventory  - external source tables and what consumes them
  * STTM              - column-level mapping with business logic
The same content is emitted as YAML for programmatic use / codegen.
"""
from __future__ import annotations

import yaml

from ..lineage import Lineage
from ..models import QvScript


_HEADERS = [
    "Layer", "Target Table", "Target Column", "Mapping Type",
    "Source Table(s)", "Source Column(s)", "Business Logic (QlikView)",
    "Snowflake SQL", "Review Notes",
]


def _src_join(pairs):
    tables, cols = [], []
    for tbl, col in pairs:
        if tbl not in tables:
            tables.append(tbl)
        cols.append(f"{tbl}.{col}")
    return " | ".join(tables), " | ".join(cols)


def _consumers(source_id: str, lin: Lineage) -> list[str]:
    node = f"source:{source_id}"
    return sorted({d for u, d in lin.table_edges if u == node})


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def write_xlsx(script: QvScript, lin: Lineage, path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    layer_fill = {
        "source": "DDEBF7", "staging": "E2EFDA",
        "intermediate": "FFF2CC", "mart": "FCE4D6", "mapping": "EDEDED",
    }

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

    # -- Sheet 1: Target Inventory -----------------------------------------
    ws = wb.active
    ws.title = "Target Inventory"
    inv_hdr = ["Table", "Layer / Role", "Load Kind", "Source / Upstream",
               "Columns", "Joins", "Review Items"]
    ws.append(inv_hdr)
    for t in script.tables:
        ws.append([
            t.name, t.layer, t.kind.value,
            (t.source or "")[:80], len(t.fields),
            ", ".join(j.right_table for j in t.joins) or "-",
            len([w for f in t.fields for w in f.warnings]) + len(t.warnings),
        ])
        ws.cell(row=ws.max_row, column=2).fill = PatternFill(
            "solid", fgColor=layer_fill.get(t.layer, "FFFFFF"))
    style_header(ws, len(inv_hdr))

    # -- Sheet 2: Source Inventory -----------------------------------------
    ws2 = wb.create_sheet("Source Inventory")
    s_hdr = ["Source Table", "Kind", "Locator", "Consumed By (# tables)",
             "Consumers"]
    ws2.append(s_hdr)
    for src in script.sources:
        cons = _consumers(src.identifier, lin)
        ws2.append([src.identifier, src.kind.value, src.locator,
                    len(cons), ", ".join(cons)])
    style_header(ws2, len(s_hdr))

    # -- Sheet 3: STTM (column mapping) ------------------------------------
    ws3 = wb.create_sheet("STTM")
    ws3.append(_HEADERS)
    for cm in lin.columns:
        stbl, scol = _src_join(cm.ultimate_sources)
        ws3.append([
            cm.layer, cm.table, cm.column, cm.mapping_type,
            stbl, scol, cm.qlik_expr, cm.snowflake_sql,
            "; ".join(cm.notes),
        ])
        ws3.cell(row=ws3.max_row, column=1).fill = PatternFill(
            "solid", fgColor=layer_fill.get(cm.layer, "FFFFFF"))
    style_header(ws3, len(_HEADERS))

    # -- Sheet 4: Legend ----------------------------------------------------
    ws4 = wb.create_sheet("Legend")
    ws4.append(["Mapping Type", "Meaning"])
    for k, v in [
        ("direct", "1:1 passthrough of a single source column"),
        ("derived", "computed from one or more columns via expression"),
        ("aggregate", "GROUP BY aggregation (SUM/COUNT/AVG/...)"),
        ("lookup", "resolved via a mapping table (ApplyMap / apply_map macro)"),
        ("constant", "literal / no source column"),
        ("join", "column supplied by a joined table"),
    ]:
        ws4.append([k, v])
    style_header(ws4, 2)

    # column widths
    widths = {
        "Target Inventory": [26, 14, 10, 46, 9, 24, 12],
        "Source Inventory": [26, 8, 60, 14, 50],
        "STTM": [12, 22, 22, 12, 26, 34, 40, 40, 30],
        "Legend": [14, 60],
    }
    for name, ws_ in [("Target Inventory", ws), ("Source Inventory", ws2),
                      ("STTM", ws3), ("Legend", ws4)]:
        for i, w in enumerate(widths[name], start=1):
            ws_.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)


# ---------------------------------------------------------------------------
# YAML
# ---------------------------------------------------------------------------

def write_yaml(script: QvScript, lin: Lineage, path: str) -> None:
    targets = []
    for t in script.tables:
        cols = []
        for cm in lin.for_table(t.name):
            cols.append({
                "name": cm.column,
                "mapping_type": cm.mapping_type,
                "sources": [{"table": a, "column": b}
                            for a, b in cm.ultimate_sources],
                "business_logic_qlikview": cm.qlik_expr,
                "snowflake_sql": cm.snowflake_sql,
                "review_notes": cm.notes,
            })
        targets.append({
            "table": t.name,
            "layer": t.layer,
            "load_kind": t.kind.value,
            "upstream": t.source,
            "columns": cols,
        })
    doc = {
        "script": script.name,
        "sources": [{"table": s.identifier, "kind": s.kind.value,
                     "locator": s.locator,
                     "consumed_by": _consumers(s.identifier, lin)}
                    for s in script.sources],
        "targets": targets,
    }
    with open(path, "w", encoding="utf-8") as fh:
        yaml.safe_dump(doc, fh, sort_keys=False, allow_unicode=True,
                       width=120)


def generate(script: QvScript, lin: Lineage, xlsx_path: str,
             yaml_path: str) -> None:
    write_xlsx(script, lin, xlsx_path)
    write_yaml(script, lin, yaml_path)
